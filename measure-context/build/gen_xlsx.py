# -*- coding: utf-8 -*-
"""Excel (.xlsx) renderer with openpyxl.

Body + tables live in cells; the two captures are embedded as real image bytes
(add_image = paste-equivalent). Diagrams are rendered as styled cell tables:
openpyxl has no worksheet autoshape/connector API, so the cell grid is the
native primitive used (reported as a limitation, not faked via raw XML).

build() returns (path, logical_text) where logical_text is every string value
written to a cell -- the (a) logical-content measure.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

import content as C
import figdata as F
from common import dwidth
from gen_text import SAMPLES, _meta_lines

_THIN = Side(style="thin", color="B8C2D4")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEAD_FILL = PatternFill("solid", fgColor="DAE8FC")
_WRAP = Alignment(wrap_text=True, vertical="top")


class _Sheet:
    def __init__(self, ws):
        self.ws = ws
        self.r = 1
        self.L = []

    def title(self, text):
        c = self.ws.cell(self.r, 1, text)
        c.font = Font(bold=True, size=16)
        self.L.append(text)
        self.r += 2

    def line(self, text, bold=False):
        c = self.ws.cell(self.r, 1, text)
        c.font = Font(bold=bold)
        self.L.append(text)
        self.r += 1

    def heading(self, text):
        self.r += 1
        c = self.ws.cell(self.r, 1, text)
        c.font = Font(bold=True, size=12)
        self.L.append(text)
        self.r += 1

    def para(self, text):
        self.ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=8)
        c = self.ws.cell(self.r, 1, text)
        c.alignment = _WRAP
        self.L.append(text)
        self.ws.row_dimensions[self.r].height = 15 * (1 + dwidth(text) // 110)
        self.r += 1

    def table(self, header, rows):
        for ci, h in enumerate(header, start=1):
            c = self.ws.cell(self.r, ci, str(h))
            c.font = Font(bold=True)
            c.fill = _HEAD_FILL
            c.border = _BORDER
            c.alignment = _WRAP
            self.L.append(str(h))
        self.r += 1
        for row in rows:
            for ci, val in enumerate(row, start=1):
                c = self.ws.cell(self.r, ci, str(val))
                c.border = _BORDER
                c.alignment = _WRAP
                self.L.append(str(val))
            self.r += 1
        self.r += 1

    def image(self, path):
        img = XLImage(path)
        self.ws.add_image(img, "A%d" % self.r)
        self.r += 38


def build():
    wb = Workbook()
    sh = _Sheet(wb.active)
    sh.ws.title = "仕様書"
    for col, w in {"A": 16, "B": 26, "C": 12, "D": 18, "E": 12, "F": 22, "G": 12, "H": 22}.items():
        sh.ws.column_dimensions[col].width = w

    sh.title(C.DOC["title"])
    sh.line(" / ".join("%s: %s" % kv for kv in _meta_lines()))
    sh.para("概要: " + C.DOC["abstract"])

    dt = F.diagram_tables()

    def section(num):
        n, head, paras = next(s for s in C.SECTIONS if s[0] == num)
        sh.heading("%s %s" % (n, head))
        for p in paras:
            sh.para(p)

    def rel(key):
        for (title, header, rows) in dt[key]:
            sh.line(title, bold=True)
            sh.table(header, rows)

    section("1")
    rel("block")
    section("2")
    section("3")
    rel("state")
    section("4")
    rel("sequence")
    section("5")
    rel("flow")
    section("6")
    section("7")
    sh.line(C.PARAM_TABLE_CAPTION, bold=True)
    sh.table(C.PARAM_TABLE_HEADER, C.PARAM_TABLE_ROWS)
    sh.line(C.IF_TABLE_CAPTION, bold=True)
    sh.table(C.IF_TABLE_HEADER, C.IF_TABLE_ROWS)

    for cap, img in [(C.FIG_CAP1, C.CAPTURE1), (C.FIG_CAP2, C.CAPTURE2)]:
        sh.line(cap, bold=True)
        sh.image(os.path.join(SAMPLES, img))

    path = os.path.join(SAMPLES, "spec_lkas.xlsx")
    wb.save(path)
    return path, "\n".join(sh.L)


if __name__ == "__main__":
    from common import count_tokens
    p, logical = build()
    print("xlsx logical:", count_tokens(logical), "tokens,", len(logical), "chars ->", os.path.basename(p))
